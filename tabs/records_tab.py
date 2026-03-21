import numpy as np
import pandas as pd
import pyqtgraph as pg
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QComboBox,
    QMessageBox,
    QFileDialog,
)

from models import CSV_PATH
from storage import load_records_df
from theme import btn_neutral, btn_delete, btn_export
from widgets.history_table import HistoryTable


class RecordsTab(QWidget):
    def __init__(self):
        super().__init__()

        root = QVBoxLayout(self)

        top = QHBoxLayout()
        root.addLayout(top)

        self.session_combo = QComboBox()
        self.btn_refresh_sessions = QPushButton("Refresh Sessions")
        self.btn_delete_session = QPushButton("Delete Session")
        self.btn_export = QPushButton("Export Excel…")

        top.addWidget(QLabel("Session:"))
        top.addWidget(self.session_combo, 4)
        top.addWidget(self.btn_refresh_sessions)
        top.addWidget(self.btn_delete_session)
        top.addWidget(self.btn_export)

        self.btn_refresh_sessions.clicked.connect(self.refresh_sessions)
        self.session_combo.currentIndexChanged.connect(self.load_selected_session_plot)
        self.btn_delete_session.clicked.connect(self.delete_selected_session)
        self.btn_export.clicked.connect(self.export_selected_session)

        # Apply button stylesheets
        self.btn_refresh_sessions.setStyleSheet(btn_neutral())
        self.btn_delete_session.setStyleSheet(btn_delete())
        self.btn_export.setStyleSheet(btn_export())

        self.session_info = QLabel("—")
        self.session_info.setStyleSheet("color: white;")
        root.addWidget(self.session_info)

        self.session_indices_info = QLabel("—")
        self.session_indices_info.setStyleSheet("color: white; font-weight:600;")
        root.addWidget(self.session_indices_info)

        # Index graph for past records
        self.plot_hist = pg.PlotWidget()
        self.plot_hist.showGrid(x=True, y=True)
        self.plot_hist.setLabel("left", "Index Value")
        self.plot_hist.setLabel("bottom", "Samples (session)")
        self.plot_hist.setBackground("#1e1650")
        root.addWidget(self.plot_hist, 1)

        self.hist_legend = self._make_legend(self.plot_hist)

        self.curves_hist = {}
        index_cfg = {
            "chlorophyll_index": ("Chlorophyll Index", "#10b981"),
            "car_chl_ratio": ("Car:Chl Ratio", "#3b82f6"),
            "yellow_index": ("Yellow Index", "#f59e0b"),
            "stress_ratio": ("Stress Ratio", "#ef4444"),
        }

        for key, (label, color) in index_cfg.items():
            pen = pg.mkPen(color=color, width=3)
            curve = pg.PlotDataItem([], [], pen=pen)
            self.plot_hist.addItem(curve)
            self.curves_hist[key] = curve
            self.hist_legend.addItem(curve, self._colored_label(label, color))

        records_header = QHBoxLayout()
        root.addLayout(records_header)

        records_title = QLabel("📋  Session History")
        records_title.setStyleSheet("font-size:18px;font-weight:800;color:white;")
        self.records_history_count = QLabel("0 snaps recorded")
        self.records_history_count.setStyleSheet("color:#c4b5fd;font-weight:600;")

        records_header.addWidget(records_title)
        records_header.addStretch()
        records_header.addWidget(self.records_history_count)

        self.records_history_table = HistoryTable()
        self.records_history_table.setMinimumHeight(180)
        root.addWidget(self.records_history_table, 1)

    def _make_legend(self, plot_widget: pg.PlotWidget):
        leg = plot_widget.addLegend(offset=(10, 10))
        if leg is None:
            leg = pg.LegendItem(offset=(10, 10))
            leg.setParentItem(plot_widget.getPlotItem())
        return leg

    def _colored_label(self, text: str, color_hex: str) -> str:
        return f'<span style="color:{color_hex}; font-weight:700;">{text}</span>'

    def _autofit_columns(self, ws):
        """Auto-fit column widths based on the longest content in each column.
        
        Args:
            ws: openpyxl worksheet to auto-fit columns for
        """
        for col_idx, col in enumerate(ws.iter_cols(), 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            
            for cell in col:
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
            
            adjusted_width = min(max_length + 4, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

    def refresh_sessions(self):
        df = load_records_df(CSV_PATH)
        if df.empty:
            self.session_combo.clear()
            self.session_combo.addItem("No sessions")
            self.session_info.setText("—")
            self.session_indices_info.setText("—")
            self.clear_hist_plot()
            self.records_history_table.setRowCount(0)
            self.records_history_count.setText("0 snaps recorded")
            return

        sessions = (
            df.groupby("session_id")
            .agg(
                start=("time", "min"),
                end=("time", "max"),
                count=("time", "count"),
                session_name=("session_name", "last"),
                overall_status=("overall_status", "last"),
            )
            .sort_values("end", ascending=False)
            .reset_index()
        )

        def label_row(r):
            nm = str(r["session_name"]).strip()
            status = str(r["overall_status"]).strip()
            if not nm:
                nm = "Unnamed"
            return f"{nm} [{status}] ({r['end']:%Y-%m-%d %I:%M %p})"

        self._sessions_df = sessions
        self.session_combo.blockSignals(True)
        self.session_combo.clear()
        for _, r in sessions.iterrows():
            self.session_combo.addItem(label_row(r), userData=str(r["session_id"]))
        self.session_combo.blockSignals(False)

        self.load_selected_session_plot()

    def clear_hist_plot(self):
        for key in self.curves_hist:
            self.curves_hist[key].setData([], [])

    def load_selected_session_plot(self):
        sid = self.session_combo.currentData()
        if not sid or str(sid).lower().startswith("no sessions"):
            self.session_info.setText("—")
            self.session_indices_info.setText("—")
            self.clear_hist_plot()
            self.records_history_table.setRowCount(0)
            self.records_history_count.setText("0 snaps recorded")
            return

        df = load_records_df(CSV_PATH)
        sdf = df[df["session_id"].astype(str) == str(sid)].sort_values("time").copy()
        if sdf.empty:
            self.session_info.setText("Session empty.")
            self.session_indices_info.setText("—")
            self.clear_hist_plot()
            self.records_history_table.setRowCount(0)
            self.records_history_count.setText("0 snaps recorded")
            return

        start = sdf["time"].iloc[0]
        end = sdf["time"].iloc[-1]
        count = len(sdf)
        name = str(sdf["session_name"].iloc[-1]).strip() or "Unnamed"
        overall = str(sdf["overall_status"].iloc[-1]).strip() or "—"

        self.session_info.setText(
            f"Session: {name} | {start:%Y-%m-%d %I:%M:%S %p} → {end:%I:%M:%S %p} | "
            f"Samples: {count} | Overall: {overall} | ID: {sid}"
        )

        try:
            last = sdf.iloc[-1]
            self.session_indices_info.setText(
                f"Chl={float(last['chlorophyll_index']):.4f} "
                f"(Δ {float(last['chlorophyll_index_delta_pct']):.2f}% | {last['chlorophyll_index_status']})   |   "
                f"Car:Chl={float(last['car_chl_ratio']):.4f} "
                f"(Δ {float(last['car_chl_ratio_delta_pct']):.2f}% | {last['car_chl_ratio_status']})   |   "
                f"Yellow={float(last['yellow_index']):.4f} "
                f"(Δ {float(last['yellow_index_delta_pct']):.2f}% | {last['yellow_index_status']})   |   "
                f"Stress={float(last['stress_ratio']):.4f} "
                f"(Δ {float(last['stress_ratio_delta_pct']):.2f}% | {last['stress_ratio_status']})"
            )
        except Exception:
            self.session_indices_info.setText("Index summary unavailable.")

        x = np.arange(count)

        for key in self.curves_hist:
            if key in sdf.columns:
                y = pd.to_numeric(sdf[key], errors="coerce").to_numpy()
                self.curves_hist[key].setData(x, y)
            else:
                self.curves_hist[key].setData([], [])

        self.records_history_table.populate_records_history_table(sdf, self.records_history_count)

    def delete_selected_session(self):
        sid = self.session_combo.currentData()
        if not sid:
            return

        reply = QMessageBox.question(
            self,
            "Delete Session",
            "Delete this entire session from the CSV? This cannot be undone.",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        df = load_records_df(CSV_PATH)
        new_df = df[df["session_id"].astype(str) != str(sid)].copy()
        new_df2 = new_df.copy()
        new_df2["time"] = new_df2["time"].dt.strftime("%Y-%m-%d %H:%M:%S.%f")
        new_df2.to_csv(CSV_PATH, index=False)

        self.refresh_sessions()

    def export_selected_session(self):
        # Get selected session ID
        sid = self.session_combo.currentData()
        if not sid:
            return

        # Load and filter DataFrame to this session
        df = load_records_df(CSV_PATH)
        sdf = df[df["session_id"].astype(str) == str(sid)].sort_values("time").copy()
        if sdf.empty:
            QMessageBox.critical(self, "Error", "Session has no data.")
            return

        # Open save dialog
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Session", "", "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        try:
            # Define columns for Session Data sheet
            session_data_cols = [
                "time",
                "F1",
                "F2",
                "F3",
                "F4",
                "F5",
                "F6",
                "F7",
                "F8",
                "CLR",
                "chlorophyll_index",
                "chlorophyll_index_delta_pct",
                "car_chl_ratio",
                "car_chl_ratio_delta_pct",
                "yellow_index",
                "yellow_index_delta_pct",
                "stress_ratio",
                "stress_ratio_delta_pct",
                "overall_status",
            ]

            # Filter to only include columns that exist in the dataframe
            available_cols = [col for col in session_data_cols if col in sdf.columns]
            session_data = sdf[available_cols].copy()

            # Create workbook and worksheets
            wb = Workbook()
            ws_session = wb.active
            ws_session.title = "Session Data"
            ws_summary = wb.create_sheet("Summary")

            # ============ Session Data Sheet ============

            # Write header row
            for col_idx, col_name in enumerate(available_cols, 1):
                cell = ws_session.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.font = Font(
                    name="Calibri", size=11, bold=True, color="FFFFFF"
                )
                cell.fill = PatternFill(start_color="281C59", end_color="281C59", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Write data rows with banded formatting and status color coding
            color_1 = "1e1650"
            color_2 = "2d2070"
            status_colors = {
                "Healthy": {"fill": "052e16", "text": "86efac"},
                "Mild stress": {"fill": "3f2a06", "text": "fde047"},
                "Stressed": {"fill": "3f1113", "text": "f87171"},
            }

            for row_idx, (_, row_data) in enumerate(session_data.iterrows(), 2):
                # Determine banded row color
                band_color = color_1 if (row_idx - 2) % 2 == 0 else color_2

                for col_idx, col_name in enumerate(available_cols, 1):
                    cell = ws_session.cell(row=row_idx, column=col_idx)
                    value = row_data[col_name]

                    # Format datetime
                    if pd.api.types.is_datetime64_any_dtype(type(value)):
                        cell.value = value
                        cell.number_format = "YYYY-MM-DD HH:MM:SS"
                    else:
                        cell.value = value

                    cell.font = Font(name="Calibri", size=10, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                    # Apply status color coding to overall_status column
                    if col_name == "overall_status":
                        status = str(value).strip()
                        if status in status_colors:
                            config = status_colors[status]
                            cell.fill = PatternFill(
                                start_color=config["fill"],
                                end_color=config["fill"],
                                fill_type="solid",
                            )
                            cell.font = Font(
                                name="Calibri", size=10, bold=True, color=config["text"]
                            )
                        else:
                            cell.fill = PatternFill(
                                start_color=band_color,
                                end_color=band_color,
                                fill_type="solid",
                            )
                    else:
                        # Apply banded row color
                        cell.fill = PatternFill(
                            start_color=band_color,
                            end_color=band_color,
                            fill_type="solid",
                        )

            # Apply conditional formatting to delta percentage columns
            delta_cols = [
                "chlorophyll_index_delta_pct",
                "car_chl_ratio_delta_pct",
                "yellow_index_delta_pct",
                "stress_ratio_delta_pct",
            ]
            for col_idx, col_name in enumerate(available_cols, 1):
                if col_name in delta_cols:
                    for row_idx in range(2, ws_session.max_row + 1):
                        cell = ws_session.cell(row=row_idx, column=col_idx)
                        value = cell.value

                        # Get current fill color to preserve it
                        current_fill = cell.fill

                        # Check if value is NaN or None
                        if value is None or (isinstance(value, float) and pd.isna(value)):
                            cell.value = "—"
                            cell.font = Font(name="Calibri", size=10, color="9ca3af")
                        else:
                            try:
                                num_value = float(value)
                                if num_value >= 0:
                                    cell.font = Font(name="Calibri", size=10, color="6ee7b7")
                                else:
                                    cell.font = Font(name="Calibri", size=10, color="f87171")
                            except (ValueError, TypeError):
                                cell.font = Font(name="Calibri", size=10, color="9ca3af")

            # Auto-fit columns in Session Data sheet
            self._autofit_columns(ws_session)

            # Freeze panes (freeze row 1)
            ws_session.freeze_panes = "A2"

            # ============ Summary Sheet ============

            # Get first and last rows
            first_row = sdf.iloc[0]
            last_row = sdf.iloc[-1]

            # Extract session info
            session_name = str(first_row.get("session_name", "Unnamed")).strip() or "Unnamed"
            start_time = first_row["time"]
            end_time = last_row["time"]
            count = len(sdf)

            # Row 1: Session name (merged A1:B1)
            ws_summary.merge_cells("A1:B1")
            cell_name = ws_summary["A1"]
            cell_name.value = session_name
            cell_name.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
            cell_name.fill = PatternFill(
                start_color="281C59", end_color="281C59", fill_type="solid"
            )
            cell_name.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Row 2: Recording info (merged A2:B2)
            ws_summary.merge_cells("A2:B2")
            cell_info = ws_summary["A2"]
            cell_info.value = f"Recorded: {start_time:%Y-%m-%d %I:%M:%S %p} → {end_time:%I:%M:%S %p}  |  {count} snapshots"
            cell_info.font = Font(name="Calibri", size=10, color="9ca3af")
            cell_info.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # Row 3: Blank spacer

            # Row 4: Baseline header
            ws_summary.cell(row=4, column=1).value = "Metric"
            ws_summary.cell(row=4, column=2).value = "Baseline Value"
            for col_idx in [1, 2]:
                cell = ws_summary.cell(row=4, column=col_idx)
                cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="281C59", end_color="281C59", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Rows 5–8: Baseline metrics
            metrics_baseline = [
                ("Chlorophyll Index", "chlorophyll_index"),
                ("CAR:CHL Ratio", "car_chl_ratio"),
                ("Yellow Index", "yellow_index"),
                ("Stress Ratio", "stress_ratio"),
            ]
            for row_offset, (metric_label, metric_col) in enumerate(metrics_baseline, 5):
                # Metric name
                cell_metric = ws_summary.cell(row=row_offset, column=1)
                cell_metric.value = metric_label
                cell_metric.font = Font(name="Calibri", size=10, color="FFFFFF")
                cell_metric.fill = PatternFill(
                    start_color=color_1, end_color=color_1, fill_type="solid"
                )
                cell_metric.alignment = Alignment(horizontal="left", vertical="center")

                # Value
                cell_value = ws_summary.cell(row=row_offset, column=2)
                if metric_col in first_row.index:
                    try:
                        cell_value.value = float(first_row[metric_col])
                        cell_value.number_format = "0.0000"
                    except (ValueError, TypeError):
                        cell_value.value = first_row[metric_col]
                else:
                    cell_value.value = "N/A"
                cell_value.font = Font(name="Calibri", size=10, color="FFFFFF")
                cell_value.fill = PatternFill(
                    start_color=color_1, end_color=color_1, fill_type="solid"
                )
                cell_value.alignment = Alignment(horizontal="center", vertical="center")

            # Row 9: Blank spacer

            # Row 10: Final header
            ws_summary.cell(row=10, column=1).value = "Metric"
            ws_summary.cell(row=10, column=2).value = "Final Value"
            ws_summary.cell(row=10, column=3).value = "Delta %"
            ws_summary.cell(row=10, column=4).value = "Status"
            for col_idx in [1, 2, 3, 4]:
                cell = ws_summary.cell(row=10, column=col_idx)
                cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="281C59", end_color="281C59", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Rows 11–14: Final metrics with color-coded status
            metrics_final = [
                ("Chlorophyll Index", "chlorophyll_index", "chlorophyll_index_delta_pct", "chlorophyll_index_status"),
                ("CAR:CHL Ratio", "car_chl_ratio", "car_chl_ratio_delta_pct", "car_chl_ratio_status"),
                ("Yellow Index", "yellow_index", "yellow_index_delta_pct", "yellow_index_status"),
                ("Stress Ratio", "stress_ratio", "stress_ratio_delta_pct", "stress_ratio_status"),
            ]
            for row_offset, (metric_label, metric_col, delta_col, status_col) in enumerate(metrics_final, 11):
                # Metric name
                cell_metric = ws_summary.cell(row=row_offset, column=1)
                cell_metric.value = metric_label
                cell_metric.font = Font(name="Calibri", size=10, color="FFFFFF")
                cell_metric.fill = PatternFill(
                    start_color=color_2, end_color=color_2, fill_type="solid"
                )
                cell_metric.alignment = Alignment(horizontal="left", vertical="center")

                # Final value
                cell_final = ws_summary.cell(row=row_offset, column=2)
                if metric_col in last_row.index:
                    try:
                        cell_final.value = float(last_row[metric_col])
                        cell_final.number_format = "0.0000"
                    except (ValueError, TypeError):
                        cell_final.value = last_row[metric_col]
                else:
                    cell_final.value = "N/A"
                cell_final.font = Font(name="Calibri", size=10, color="FFFFFF")
                cell_final.fill = PatternFill(
                    start_color=color_2, end_color=color_2, fill_type="solid"
                )
                cell_final.alignment = Alignment(horizontal="center", vertical="center")

                # Delta %
                cell_delta = ws_summary.cell(row=row_offset, column=3)
                if delta_col in last_row.index:
                    try:
                        cell_delta.value = float(last_row[delta_col])
                        cell_delta.number_format = "0.00"
                    except (ValueError, TypeError):
                        cell_delta.value = last_row[delta_col]
                else:
                    cell_delta.value = "N/A"
                cell_delta.font = Font(name="Calibri", size=10, color="FFFFFF")
                cell_delta.fill = PatternFill(
                    start_color=color_2, end_color=color_2, fill_type="solid"
                )
                cell_delta.alignment = Alignment(horizontal="center", vertical="center")

                # Status with color coding
                cell_status = ws_summary.cell(row=row_offset, column=4)
                status_text = str(last_row.get(status_col, "Unknown")).strip() if status_col in last_row.index else "Unknown"
                cell_status.value = status_text
                cell_status.font = Font(name="Calibri", size=10, color="FFFFFF", bold=True)
                cell_status.alignment = Alignment(horizontal="center", vertical="center")

                # Map status names to colors (normalize status text)
                status_normalized = status_text.lower()
                if "healthy" in status_normalized:
                    cell_status.fill = PatternFill(
                        start_color="052e16", end_color="052e16", fill_type="solid"
                    )
                    cell_status.font = Font(name="Calibri", size=10, bold=True, color="86efac")
                elif "mild" in status_normalized or "mild stress" in status_normalized:
                    cell_status.fill = PatternFill(
                        start_color="3f2a06", end_color="3f2a06", fill_type="solid"
                    )
                    cell_status.font = Font(name="Calibri", size=10, bold=True, color="fde047")
                elif "stress" in status_normalized:
                    cell_status.fill = PatternFill(
                        start_color="3f1113", end_color="3f1113", fill_type="solid"
                    )
                    cell_status.font = Font(name="Calibri", size=10, bold=True, color="f87171")
                else:
                    cell_status.fill = PatternFill(
                        start_color=color_2, end_color=color_2, fill_type="solid"
                    )

            # Auto-fit columns in Summary sheet
            self._autofit_columns(ws_summary)

            # Set row heights for merged cells
            ws_summary.row_dimensions[1].height = 25
            ws_summary.row_dimensions[2].height = 20

            # Save the workbook
            wb.save(file_path)

            QMessageBox.information(
                self,
                "Export Successful",
                f"Saved to:\n{file_path}",
            )
        except Exception as e:
            QMessageBox.critical(
                self, "Export Failed", f"Failed to export:\n{str(e)}"
            )

